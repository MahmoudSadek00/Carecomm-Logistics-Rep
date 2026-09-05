"""
Engine for the Shipping + Shopify consolidation tool.

Built from REAL samples (Aug 2026):
  - Shopify side: a "Monthly POS Report" / Orders-by-staff-member style Shopify
    Analytics export (Day, Sales channel, Shipping country, Order name,
    Staff member name, Total sales, Net sales, Orders (first-time),
    Orders (returning), ...).
  - Shipping side: TWO different carrier formats seen so far --
      * Gulf: "Golden Collection" (Reference, Waybill, Carrier Waybill,
        Consignee Contact, Consignee City, ...)
      * Iraq: a different export entirely (Order #, Recipient Name,
        Recipient Phone, City, Subtotal, Total, ...) -- no waybill/AWB field
        at all, and it DOES carry a consignee/recipient name, which Iraq's
        raw sheet template needs and UAE/Gulf's doesn't.
    UAE & Oman's own shipping export hasn't been sampled yet -- assumed to
    match the Gulf "Golden Collection" format since it's the same carrier;
    confirm on the first real file.
  - Target layout: sheet_templates.xlsx (the ops team's own template), one
    tab per raw Google Sheet (UAE & Oman / Gulf / Iraq).

Scope, on purpose: this tool only fills the columns that are already known
the moment an order is created and a shipping label is generated. It does
NOT fill "shipping date" / "Analysis" (delivery status) / "Notes" -- those
only exist once the order has actually moved, and stay part of the ops
team's existing manual process (or a later sync-script integration).

--- Order Value: the multi-row-per-order fix (Aug 2026) ------------------
A "Monthly POS Report"-style Shopify export has ONE row per order, PLUS one
extra row for every LATER edit against that same Order name -- either a
partial/full return (a row that's the exact negative of the original) or an
item added afterward (a row with its own positive Total sales). Confirmed
against the real Iraq export: in every "item added" case found, that later
row has Total sales == Net sales (no extra shipping gets charged just for
adding an item to an already-created order) -- so the order's real shipping
fee always stays anchored to the original row.

The correct Order Value is SUM(Total sales) across every row sharing that
Order name, not just the first row (that would silently drop a later
addition) and not "drop every Orders=0 row" (an earlier version of this tool
did that -- WRONG, it silently dropped the added-item revenue along with the
harmless return-adjustment rows). Summation is linear, so
SUM(Total) - SUM(Net) across the group equals summing the per-row
(Total - Net) difference -- there's no discrepancy to worry about between
"sum then subtract" and "subtract then sum".

All other per-order fields (date, city, salesman, new/returning-customer)
are taken from the EARLIEST row in the group (the original order-creation
row) -- a later edit-row doesn't change who ordered it or when.
"""
import io
import re
import unicodedata
import datetime as dt

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Per-country output layout. Each field's 'source' says which uploaded file
# it comes from. 'ref_number' is always present on both sides -- it's the
# join key. Order here = column order in the downloaded file.
# 'ref_prefix': whether the output Reference/Receipt column keeps the
# leading '#' -- UAE/Gulf templates show "#90127", Iraq's shows "IQ-1059"
# (no '#'), even though Shopify and the shipping file both carry the '#' on
# every source. Matching itself always ignores '#' either way (see clean_key).
# ---------------------------------------------------------------------------
TARGET_SHEETS = {
    'uae_om': {
        'label': 'UAE & Oman',
        'ref_prefix': True,
        # Switched to shipping-first (Aug 2026, per Mahmoud) -- same as Gulf
        # and Iraq now: an order only shows up once it's actually with the
        # shipping company, not the moment it's placed in Shopify. All three
        # groups are joined the same way now; UAE & Oman used to be the one
        # exception (Shopify-first, still in git history if ever needed).
        'join_base': 'shipping',
        # Full raw-sheet layout (confirmed against sheet_templates.xlsx's own
        # header row, Aug 2026): 'shipping date' / 'date' / the unnamed
        # column / 'Notes' / 'Analysis' are the real sheet's own blank
        # columns -- delivery outcome isn't known yet at this stage, same as
        # Iraq. 'Shipping' is baked in right after 'Order Value' (per
        # Mahmoud's spec), same derivation as Iraq: Total sales - Net sales.
        # Primary (and only) Shopify source is the "Monthly POS Report"
        # export, same as Iraq. Order Value = Net sales x 1.05 (adding back
        # 5% VAT -- UAE & Oman's own standard rate) -- see the 'uae_vat'
        # source below and the README's "UAE & Oman Order Value: Net x 1.05"
        # section for why (confirmed Sep 2026, per Mahmoud: this reconstructs
        # the real goods-only, tax-inclusive value at ~98.8% accuracy against
        # a real order-level export). Shipping is then Total sales minus that
        # same Net x1.05 figure, not raw Net -- keeps Value + Shipping =
        # Total sales, same principle as Iraq's plain Total-minus-Net, just
        # adjusted for the VAT add-back. Mahmoud confirmed (Sep 2026) this
        # VAT gap is UAE & Oman-specific -- Iraq's and Gulf's own Net sales
        # columns are already correct as-is, no adjustment needed there.
        # An earlier version of this tool (Sep 2026) also supported an
        # optional second "Export orders" upload to override these values
        # with real Subtotal/Shipping fields -- removed per Mahmoud, back to
        # Monthly POS Report only (still in git history if ever needed).
        'fields': [
            ('shipping date', 'blank', 'blank'),
            ('Reference Number', 'ref_number', 'shipping'),
            # Sep 2026, per Mahmoud -- "Sep 2026" style label, 3rd column,
            # same across all 3 groups. See format_month_year().
            ('Month', 'month', 'month'),
            ('Order date', 'order_date', 'shopify'),
            ('Consignee City', 'city', 'shopify'),
            ('Consignee Phone', 'phone', 'shipping'),
            ('Carrier WayBill', 'waybill', 'shipping'),
            ('Salesman', 'salesman', 'shopify'),
            # Sep 2026, per Mahmoud -- UAE & Oman-specific (5% VAT add-back,
            # see the top-of-dict comment and the README). NOT the same as
            # Iraq's plain 'shopify'/'computed' below -- Iraq's own Net sales
            # column needs no VAT adjustment.
            ('Order Value', 'order_value', 'uae_vat'),
            ('Shipping', 'shipping_fee', 'uae_vat_computed'),
            ('Payment Method', 'payment_type', 'shipping'),
            ('date', 'blank', 'blank'),
            ('', 'blank', 'blank'),  # unnamed column in the real sheet
            ('Notes', 'blank', 'blank'),
            ('Analysis', 'blank', 'blank'),
            ('New Customer Orders', 'new_customer', 'shopify'),
            ('Returning Customer Orders', 'returning_customer', 'shopify'),
        ],
        'country_choices': {'UAE': 'UAE', 'OM': 'Oman'},
    },
    'gulf': {
        'label': 'Gulf (SA/QA/KW)',
        'ref_prefix': True,
        # Switched to shipping-first (Aug 2026, per Mahmoud) -- same concept
        # as Iraq: an order only shows up once it's actually with the
        # shipping company, not the moment it's placed in Shopify.
        'join_base': 'shipping',
        # Gulf's own Shopify export is a DIFFERENT report type than UAE/Iraq's
        # "Monthly POS Report" -- it's a "Sales overview" export with no Net
        # sales/Returns columns at all, and New/Returning customer comes as
        # ONE text column ("New or returning customer" = "New"/"Returning"),
        # not two separate 0/1 columns. See new_returning_combined below.
        'new_returning_combined': True,
        # Same layout as UAE & Oman, with the real sheet's own differences:
        # 'AWB' instead of 'Carrier WayBill', no 'Notes' column, and TWO
        # unnamed blank columns instead of one (confirmed against
        # sheet_templates.xlsx's own header row, Aug 2026).
        'fields': [
            ('shipping date', 'blank', 'blank'),
            ('Reference Number', 'ref_number', 'shipping'),
            # Sep 2026, per Mahmoud -- "Sep 2026" style label, 3rd column,
            # same across all 3 groups. See format_month_year().
            ('Month', 'month', 'month'),
            ('Order date', 'order_date', 'shopify'),
            # Despite its name, the raw sheet's 'Consignee City' column
            # actually holds the COUNTRY code (SA/KW/QA), not a real city --
            # confirmed by Mahmoud (Aug 2026). Sourced from the shipping
            # file's own 'Consignee Country' column (which already comes as
            # exactly 'SA'/'KW'/'QA' in the real export), normalized the same
            # way as UAE & Oman's country-code city. See 'shipping_country'
            # in fill_fields() below -- NOT the actual city text.
            ('Consignee City', 'city', 'shipping_country'),
            ('Consignee Phone', 'phone', 'shipping'),
            ('AWB', 'waybill', 'shipping'),
            ('Salesman', 'salesman', 'shopify'),
            # Order Value = Shopify's own Total sales (Sep 2026, per Mahmoud,
            # ALL Gulf countries, not just Saudi) -- SUPERSEDES the Aug 2026
            # v2 rule that used the shipping file's COD Amount/Cargo Value
            # (see 'gulf_value_full' in git history if ever needed). Reason:
            # the shipping company doesn't reliably write the real order
            # price -- for Prepaid orders especially, nothing is collected on
            # delivery so there's no reason for them to record an accurate
            # amount there at all. Reuses the plain 'shopify' path (same as
            # Iraq) -- Gulf's Sales overview format has no Net sales column,
            # so it always falls through to Total sales directly. Shipping
            # stays 0 (unchanged, still no reliable way to split it out).
            ('Order Value', 'order_value', 'shopify'),
            ('Shipping', 'shipping_fee', 'zero'),
            ('Payment Method', 'payment_type', 'shipping'),
            ('date', 'blank', 'blank'),
            ('', 'blank', 'blank'),  # unnamed column #1 in the real sheet
            ('', 'blank', 'blank'),  # unnamed column #2 in the real sheet
            ('Analysis', 'blank', 'blank'),
            ('New Customer Orders', 'new_customer', 'shopify'),
            ('Returning Customer Orders', 'returning_customer', 'shopify'),
        ],
        # Gulf's 'Consignee City' output column is actually the country code
        # (see the 'fields' comment above) -- these are the three the real
        # shipping export uses (Consignee Country: SA/KW/QA), same shape as
        # UAE & Oman's country_choices.
        'country_choices': {'SA': 'Saudi Arabia', 'KW': 'Kuwait', 'QA': 'Qatar'},
        # Shipping-side columns needed by custom 'source' branches above but
        # not tied to any single output field on their own -- forced into
        # the shipping mapping UI in app.py regardless of whether a 'fields'
        # entry declares them as its source. 'country' feeds Consignee City
        # (via 'shipping_country' above); 'payment_type' feeds the visible
        # Payment Method output column. cargo_value/cod_amount are no longer
        # needed here (Sep 2026 -- Order Value now comes from Shopify, see
        # the 'fields' comment above).
        'extra_shipping_fields': ['country', 'payment_type'],
        # Gulf's "Sales overview" export has one row per (order, day) across
        # the WHOLE report date range (all zeros except the real event day)
        # -- NOT one row per return/edit event like Monthly POS Report. Without
        # this, aggregate_shopify_orders' "earliest row" pick grabs the first
        # calendar day of the report instead of the order's actual date.
        # Forced into the Shopify mapping UI in app.py, same mechanism as
        # extra_shipping_fields above.
        'extra_shopify_fields': ['event_time'],
    },
    'iraq': {
        'label': 'Iraq',
        'ref_prefix': False,
        # All three groups are now joined this way (Aug 2026, per Mahmoud):
        # the shipping file is the base/driving table and Shopify data is
        # looked up onto it -- an order isn't logged in a raw sheet until
        # it's actually with the shipping company, so it doesn't belong in
        # the download yet either.
        'join_base': 'shipping',
        # Full raw-sheet column layout, in the sheet's own order. date ship /
        # status / Notes / Analysis are left blank on purpose -- delivery
        # outcome isn't known at this stage, same as UAE/Gulf. Shipping is
        # baked in right after Value (per Mahmoud's spec), computed as
        # Total sales - Net sales, same derivation as the optional Shipping
        # column on UAE/Gulf. Month (3rd column) and Salesman (right after
        # City) added Sep 2026, per Mahmoud -- Iraq didn't have either before
        # (unlike UAE & Oman / Gulf, which already had Salesman).
        'fields': [
            ('date ship', 'blank', 'blank'),
            ('ReceiptNumber', 'ref_number', 'shipping'),
            ('Month', 'month', 'month'),
            ('date greeting', 'order_date', 'shopify'),
            ('Name', 'consignee_name', 'shipping'),
            ('PhoneNumber', 'phone', 'shipping'),
            ('City', 'city', 'shipping'),
            ('Salesman', 'salesman', 'shopify'),
            ('Value', 'order_value', 'shopify'),
            ('Shipping', 'shipping_fee', 'computed'),
            ('Payment Method', 'payment_type', 'shipping'),
            ('status', 'blank', 'blank'),
            ('Notes', 'blank', 'blank'),
            ('Analysis', 'blank', 'blank'),
            ('New Customer Orders', 'new_customer', 'shopify'),
            ('Returning Customer Orders', 'returning_customer', 'shopify'),
        ],
        'country_choices': {},  # Iraq's City is a real city, passed through as-is, not normalized
    },
}

FIELD_LABELS = {
    'ref_number': 'Reference / Order Number (join key)',
    'order_date': 'Order date',
    'city': 'City / shipping country',
    'phone': 'Consignee phone',
    'waybill': 'Carrier waybill / AWB',
    'salesman': 'Staff / salesman name',
    'order_value': 'Order value -- goods only, Net sales (summed across an order\'s rows); Total sales used only if no Net sales column is mapped. UAE & Oman only: Net sales x 1.05 (VAT add-back) -- see the README.',
    'shipping_fee': 'Shipping fee (optional -- Total sales minus Net sales, summed across an order\'s rows; UAE & Oman uses Net x1.05 in that subtraction instead of raw Net -- see the README)',
    'new_customer': 'New-customer flag',
    'returning_customer': 'Returning-customer flag',
    'consignee_name': 'Consignee / recipient name',
    'net_sales': 'Net sales (for the optional Shipping column)',
    'new_or_returning': "New or returning customer (single combined column, Gulf's own Shopify export)",
    'country': "Consignee Country (SA/KW/QA on the shipping file -- feeds the output 'Consignee City' column, which despite its name holds the country code, not an actual city)",
    'cargo_value': "Cargo Value (goods value on the shipping file -- Gulf's Order Value source for Prepaid orders, and fallback for COD orders)",
    'cod_amount': "COD Amount (cash collected on delivery -- Gulf's primary Order Value source for COD orders; 0 for Prepaid orders by design)",
    'payment_type': 'Payment Type (COD / Prepaid)',
    'event_time': (
        "Event time column (Gulf's 'Sales overview' export only) -- this report has one row per "
        "(order, day) across the WHOLE date range, all zeros except the day something actually "
        "happened. This column is only populated on that real row, so it's used to find the "
        "correct order date/salesman/customer-type instead of the report's first calendar day."
    ),
}

# Confirmed against the real "Monthly POS Report" sample (Aug 2026, UAE + Iraq).
SHOPIFY_DEFAULTS = {
    'ref_number': ['Order name', 'Name'],
    'order_date': ['Day', 'Created at'],
    'city': ['Shipping country', 'Shipping Province Name'],
    'salesman': ['Staff member name'],
    'order_value': ['Total sales', 'Total'],
    'net_sales': ['Net sales'],
    'new_customer': ['Orders (first-time)', 'New Customer Orders'],
    'returning_customer': ['Orders (returning)', 'Returning Customer Orders'],
    # Gulf's "Sales overview" export only -- see new_returning_combined.
    'new_or_returning': ['New or returning customer'],
    # Gulf's "Sales overview" export only -- see aggregate_shopify_orders.
    'event_time': ['Hour'],
}

# Union of both real shipping-carrier formats seen so far (Gulf "Golden
# Collection" + Iraq's own export) -- guess_column only matches candidates
# that are actually present in the uploaded file, so listing both is safe.
SHIPPING_DEFAULTS = {
    'ref_number': ['Reference', 'Reference Number', 'Order #', 'Order Number'],
    'phone': ['Consignee Contact', 'Consignee Phone', 'Recipient Phone', 'Customer Phone (E.164)'],
    'waybill': ['Carrier Waybill', 'AWB'],
    'consignee_name': ['Recipient Name', 'Consignee Name', 'Customer Name'],
    'city': ['City', 'Consignee City'],
    # Gulf ("Golden Collection" export) only -- see TARGET_SHEETS['gulf'].
    'country': ['Consignee Country'],
    'cargo_value': ['Cargo Value'],
    'cod_amount': ['COD Amount'],
    'payment_type': ['Payment Type'],
}

# Shopify's 'Shipping country' is blank for every non-Online-Store channel
# (Point of Sale, Draft Orders, Mobile App, ...) since there's no shipping
# address on an in-store sale -- those rows need an explicit default picked
# in the app, they are not silently guessed here. (Iraq's city comes from
# the shipping file instead, as a real city name, so it isn't normalized
# here at all -- see TARGET_SHEETS['iraq'].)
COUNTRY_NORMALIZE = {
    'uae_om': {
        'united arab emirates': 'UAE', 'uae': 'UAE', 'ae': 'UAE',
        'oman': 'OM', 'om': 'OM',
    },
    'gulf': {
        'kuwait': 'KW', 'kw': 'KW',
        'qatar': 'QA', 'qa': 'QA',
        'saudi arabia': 'SA', 'saudi': 'SA', 'sa': 'SA', 'ksa': 'SA',
    },
}


def clean_display(s):
    """Strip invisible/combining Unicode characters and extra whitespace but
    keep original case/formatting -- used for values actually WRITTEN into
    the output, so a hidden character on either input file never gets
    carried forward (the same bug class fixed in sync_orders.py for order
    #116339)."""
    if s is None:
        return ''
    s = str(s)
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch) and (ch.isprintable() or ch == ' '))
    return re.sub(r'\s+', ' ', s).strip()


def clean_key(s):
    """Hidden-character-safe key for MATCHING two reference numbers -- same
    normalization as sync_orders.py's duplicate-row fix, uppercased and with
    a leading '#' stripped, so '#90127', '90127', and 'IQ-1059' vs
    '#IQ-1059' all match regardless of which side happens to carry the '#'
    or a stray space / hidden character."""
    d = clean_display(s).upper()
    if d.startswith('#'):
        d = d[1:]
    return d


def read_any(file) -> pd.DataFrame:
    name = getattr(file, 'name', str(file)).lower()
    if name.endswith('.csv'):
        return pd.read_csv(file, dtype=str, keep_default_na=False)
    return pd.read_excel(file, dtype=str, keep_default_na=False)


def guess_column(columns, candidates):
    lower_map = {str(c).lower().strip(): c for c in columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    return None


def default_mapping(columns, defaults):
    return {field: guess_column(columns, cands) for field, cands in defaults.items()}


def resolve_salesman(raw):
    s = clean_display(raw)
    return s if s else 'Created by customer'


def _to_number(raw):
    """Safe numeric parse for a shipping-file cell (Cargo Value / COD Amount) --
    returns None (not 0) for blank/unparseable, so callers can tell 'genuinely
    zero' apart from 'missing' where that distinction matters."""
    if raw in (None, ''):
        return None
    try:
        return float(str(raw).replace(',', '').strip())
    except (TypeError, ValueError):
        return None


def normalize_city(raw, target_key, default_for_blank):
    """Returns (value, was_blank, was_unrecognized). Blank input (POS/Draft/
    Mobile orders with no shipping address) uses default_for_blank and is
    flagged so the caller can warn with a count rather than silently guess."""
    s = clean_display(raw)
    if not s:
        return default_for_blank, True, False
    key = s.lower()
    mapped = COUNTRY_NORMALIZE.get(target_key, {}).get(key)
    if mapped:
        return mapped, False, False
    upper = s.upper()
    if upper in TARGET_SHEETS.get(target_key, {}).get('country_choices', {}):
        return upper, False, False
    return s, False, True  # pass through as-is, flagged for manual review


def parse_date_cell(value, convention='month_first'):
    """convention: 'day_first' or 'month_first' -- only matters for ambiguous
    numeric dates like 3/5/2026. ISO-style (2026-05-03) parses unambiguously
    regardless."""
    if value in (None, ''):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if not s:
        return None
    m = re.match(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})', s)
    if not m:
        try:
            return pd.to_datetime(s).date()
        except Exception:
            return None
    a, b, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if year < 100:
        year += 2000
    day, month = (a, b) if convention == 'day_first' else (b, a)
    if month > 12 and day <= 12:
        # naive read is impossible (e.g. month_first on '25/1/2026' -> month=25) --
        # 25 can only be a day, so swap rather than discard the row.
        day, month = month, day
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


# Fixed English abbreviations (Sep 2026, per Mahmoud) -- NOT date.strftime('%b'),
# which is locale-dependent (would print the server locale's own month names, e.g.
# Arabic, if that's ever set) and NOT written as a real date cell either, unlike Order
# date -- Month is a plain "Sep 2026"-style label, not a value anyone recomputes from.
_MONTH_ABBR = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def format_month_year(d):
    """'Sep 2026' style label from a date (or None if d is None/falsy) -- see
    _MONTH_ABBR above for why this doesn't just use strftime('%b')."""
    if not d:
        return None
    return f"{_MONTH_ABBR[d.month - 1]} {d.year}"


def aggregate_shopify_orders(
    shopify_df, shopify_map,
    date_convention='month_first',
    exclude_canceled=True, canceled_col=None,
    allowed_fulfillment_statuses=None, fulfillment_col=None,
):
    """Collapse a Monthly-POS-Report-style export to one row per order. See
    module docstring for why SUM(Total sales) [and SUM(Net sales), when a
    net-sales column is mapped] is the correct way to get Order Value (and
    the derived Shipping amount) across an order's rows, and why every other
    field is taken from the earliest row.

    Returns (agg_df, stats) where agg_df has one row per order with columns:
    _key, _ref_raw, _order_date, _order_value, _net_sales (or None),
    _row_count, plus the first-row's raw values for city/salesman/new/
    returning-customer under their shopify_map column names.
    """
    stats = {}
    work = shopify_df.copy()
    stats['shopify_rows_total'] = len(work)

    if canceled_col and exclude_canceled:
        is_canceled = work[canceled_col].astype(str).str.strip().str.upper().isin(['TRUE', '1', 'YES'])
        stats['canceled_excluded'] = int(is_canceled.sum())
        work = work[~is_canceled]
    else:
        stats['canceled_excluded'] = 0

    if fulfillment_col and allowed_fulfillment_statuses:
        keep = work[fulfillment_col].astype(str).str.strip().str.lower().isin(
            [s.lower() for s in allowed_fulfillment_statuses]
        )
        stats['fulfillment_excluded'] = int((~keep).sum())
        work = work[keep]
    else:
        stats['fulfillment_excluded'] = 0

    ref_col = shopify_map.get('ref_number')
    day_col = shopify_map.get('order_date')
    total_col = shopify_map.get('order_value')
    net_col = shopify_map.get('net_sales')
    # Gulf's "Sales overview" export only: one row per (order, day) across the
    # WHOLE report date range, all zeros/blank except the day something
    # actually happened -- this column is populated only on that real row.
    # Without it, "earliest row" below would grab the report's first calendar
    # day instead of the order's actual date. See FIELD_LABELS['event_time'].
    event_col = shopify_map.get('event_time')

    if not ref_col:
        stats['orders_total'] = 0
        stats['orders_with_multiple_rows'] = 0
        stats['orders_with_uncertain_date'] = 0
        return pd.DataFrame(), stats

    work = work[work[ref_col].astype(str).str.strip() != '']
    work['_key'] = work[ref_col].map(clean_key)
    work = work[work['_key'] != '']
    work['_total_num'] = pd.to_numeric(work[total_col], errors='coerce').fillna(0) if total_col else 0.0
    if net_col:
        work['_net_num'] = pd.to_numeric(work[net_col], errors='coerce').fillna(0)
    work['_day_parsed'] = work[day_col].map(lambda v: parse_date_cell(v, date_convention)) if day_col else None
    if event_col:
        # notna() first -- astype(str) alone turns an actual blank/NaN cell
        # into the literal string 'nan', which would never match '' and make
        # this check always True (silently defeating the whole point).
        work['_has_event'] = work[event_col].notna() & (work[event_col].astype(str).str.strip() != '')

    rows = []
    multi_row_orders = 0
    for key, grp in work.groupby('_key', sort=False):
        grp_sorted = grp.sort_values('_day_parsed', na_position='last', kind='stable')
        if event_col:
            event_rows = grp_sorted[grp_sorted['_has_event']]
            # Every real order should have at least one event row; if a group
            # somehow has none (e.g. a stray padding-only row), fall back to
            # the full group rather than crashing on an empty .iloc[0].
            pick_from = event_rows if len(event_rows) > 0 else grp_sorted
            real_row_count = len(event_rows) if len(event_rows) > 0 else len(grp_sorted)
        else:
            pick_from = grp_sorted
            real_row_count = len(grp)
        first = pick_from.iloc[0]
        # "Earliest row" is only the TRUE order date if the original sale row
        # is actually present in this upload -- if the file's date window
        # starts after the order was really placed, the earliest row we can
        # see is a later event instead (a return/cancellation adjustment,
        # which is negative or zero), and _order_date silently comes out
        # wrong (a real, later date, so it doesn't even look obviously off).
        # A non-positive Total sales on the picked row is the tell: a real
        # original order-placement row is always a positive sale. (Sep 2026,
        # per Mahmoud -- surfaced after the ShopifyQL "Day" column turned out
        # to mean "day of the financial event", not "day the order was
        # created", for the exact same underlying reason.)
        date_uncertain = bool(first['_total_num'] <= 0)
        row = {
            '_key': key,
            '_ref_raw': first[ref_col],
            '_order_date': first['_day_parsed'],
            '_order_value': grp['_total_num'].sum(),
            '_net_sales': grp['_net_num'].sum() if net_col else None,
            '_row_count': real_row_count,
            '_date_uncertain': date_uncertain,
        }
        for f in ('city', 'salesman', 'new_customer', 'returning_customer', 'new_or_returning'):
            col = shopify_map.get(f)
            row[f'_first_{f}'] = first.get(col, '') if col else ''
        rows.append(row)
        if real_row_count > 1:
            multi_row_orders += 1

    stats['orders_total'] = len(rows)
    stats['orders_with_multiple_rows'] = multi_row_orders
    stats['orders_with_uncertain_date'] = sum(1 for r in rows if r['_date_uncertain'])
    return pd.DataFrame(rows), stats


def merge_sources(
    target_key,
    shopify_df, shopify_map,
    shipping_df, shipping_map,
    shopify_date_convention='month_first',
    default_city_for_blank='UAE',
    exclude_canceled=True,
    canceled_col=None,
    allowed_fulfillment_statuses=None,
    fulfillment_col=None,
    include_shipping_fee=False,
):
    """Returns (merged_df, warnings, stats). One row per order, joined either
    Shopify-driven (every Shopify order shows up, shipping data filled in
    where matched) or shipping-driven (per Mahmoud, Aug 2026: only orders
    already with the shipping company show up, Shopify data filled in where
    matched -- now all three groups, UAE & Oman, Gulf, and Iraq) -- see
    TARGET_SHEETS[target_key]['join_base']. QA columns '_matched' / '_issues'
    are for the in-app preview only, dropped before the final download."""
    warnings = []
    fields = list(TARGET_SHEETS[target_key]['fields'])
    ref_prefix = TARGET_SHEETS[target_key]['ref_prefix']
    join_base = TARGET_SHEETS[target_key].get('join_base', 'shopify')
    if include_shipping_fee and not any(f == 'shipping_fee' for _, f, _ in fields):
        fields = fields + [('Shipping', 'shipping_fee', 'computed')]
    ref_header = next(h for h, f, s in fields if f == 'ref_number')

    agg, stats = aggregate_shopify_orders(
        shopify_df, shopify_map,
        date_convention=shopify_date_convention,
        exclude_canceled=exclude_canceled, canceled_col=canceled_col,
        allowed_fulfillment_statuses=allowed_fulfillment_statuses, fulfillment_col=fulfillment_col,
    )
    if stats.get('orders_with_multiple_rows'):
        warnings.append(
            f"{stats['orders_with_multiple_rows']} order(s) had more than one row in the Shopify file "
            f"(a later return or an added item) -- Order Value was SUMMED across that order's rows, not "
            f"just taken from the first row. See the README for why this is correct."
        )

    agg_index = {row['_key']: row for _, row in agg.iterrows()}
    if stats.get('orders_with_uncertain_date'):
        warnings.append(
            f"{stats['orders_with_uncertain_date']} order(s)' date might be WRONG: the earliest row seen for "
            f"that order in this Shopify file is not a positive sale (Total sales <= 0) -- which usually means "
            f"the order was actually placed BEFORE this file's date range, and only a later return/cancellation "
            f"is showing up here. Re-export the Shopify file with an earlier start date to fix this (see the "
            f"README's 'Order date can come out wrong' section) -- flagged rows are marked in '_issues'."
        )
    # The Total-sales-minus-Net-sales derivation only applies to the 'computed'
    # source (Iraq) and 'uae_vat_computed' (UAE & Oman, same idea with the VAT
    # add-back folded in -- see fill_fields()). Gulf's Shipping source is
    # 'zero' (always 0, no Net sales needed) -- warning about a missing Net
    # sales mapping there would be actively misleading. Check the sheet's
    # actual shipping_fee source directly rather than the caller's raw
    # include_shipping_fee flag -- that flag alone used to short-circuit this
    # check via `or`, firing the warning even when the real source didn't need
    # Net sales at all.
    wants_net_sales_shipping = any(f == 'shipping_fee' and s in ('computed', 'uae_vat_computed') for _, f, s in fields)
    if wants_net_sales_shipping and not shopify_map.get('net_sales'):
        warnings.append("A Shipping column is included but no Net sales column was mapped on the Shopify file -- Shipping was left blank.")

    def build_index(df, col):
        idx = {}
        if not col:
            return idx
        for i, row in df.iterrows():
            key = clean_key(row.get(col, ''))
            if key:
                idx.setdefault(key, []).append(i)
        return idx

    ship_ref_col = shipping_map.get('ref_number')
    ship_index = build_index(shipping_df, ship_ref_col)
    blank_city_count = 0
    unrecognized_city_count = 0

    def make_ref_display(ref_raw):
        d = clean_display(ref_raw)
        if not ref_prefix and d.startswith('#'):
            d = d[1:]
        elif ref_prefix and not d.startswith('#'):
            d = '#' + d
        return d

    def fill_fields(row_out, issues, orow, srow_ship):
        """orow: the aggregated Shopify row (dict-like) or None if unmatched.
        srow_ship: the raw shipping row (Series) or None if unmatched."""
        nonlocal blank_city_count, unrecognized_city_count
        for header, field, source in fields:
            if field == 'ref_number':
                continue
            if field == 'blank':
                row_out[header] = None
                continue
            if source == 'shopify':
                if orow is None:
                    val = None
                elif field == 'order_date':
                    val = orow['_order_date']
                elif field == 'city':
                    val, was_blank, was_unrec = normalize_city(orow['_first_city'], target_key, default_city_for_blank)
                    if was_blank:
                        blank_city_count += 1
                    if was_unrec:
                        unrecognized_city_count += 1
                        issues.append(f"unrecognized shipping country: {orow['_first_city']!r}")
                elif field == 'salesman':
                    val = resolve_salesman(orow['_first_salesman'])
                elif field == 'order_value':
                    # Total sales INCLUDES shipping (subtotal + taxes + fees +
                    # shipping + reversals, per Shopify's own definition) --
                    # once Shipping is broken out as its own column, Value
                    # should be the goods-only amount (Net sales), or
                    # Value + Shipping would double-count the shipping fee.
                    # Falls back to Total sales when no Net sales column was
                    # mapped at all (nothing to split it with) -- which is
                    # ALWAYS the case for Gulf (its Sales overview format has
                    # no Net sales column at all, see TARGET_SHEETS['gulf']),
                    # so this always resolves to plain Total sales for Gulf.
                    # UAE & Oman uses 'uae_vat' below instead (VAT add-back).
                    val = orow['_net_sales'] if orow['_net_sales'] is not None else orow['_order_value']
                elif field in ('new_customer', 'returning_customer'):
                    if TARGET_SHEETS[target_key].get('new_returning_combined'):
                        # Gulf's Shopify export has ONE text column ("New" /
                        # "Returning") instead of two separate 0/1 columns --
                        # split it back into the two output columns here.
                        raw = clean_display(orow['_first_new_or_returning']).strip().lower()
                        want = 'new' if field == 'new_customer' else 'returning'
                        val = 1 if raw == want else 0
                    else:
                        val = clean_display(orow[f'_first_{field}'])
                else:
                    val = ''
            elif source == 'computed':
                # Iraq only (see 'uae_vat_computed' below for UAE & Oman).
                val = (orow['_order_value'] - orow['_net_sales']) if (orow is not None and orow['_net_sales'] is not None) else None
            elif source == 'uae_vat':
                # UAE & Oman Order Value (Sep 2026, per Mahmoud). Net sales
                # excludes both VAT and shipping; the real goods-only,
                # tax-inclusive value (what Shopify calls Subtotal) is
                # Net sales x 1.05 -- UAE & Oman's standard 5% VAT rate,
                # confirmed at ~98.8% accuracy against a real order-level
                # export (see the README's "UAE & Oman Order Value: Net x
                # 1.05" section for the worked examples and the ~1.2% gap's
                # cause). Iraq and Gulf do NOT get this adjustment -- Mahmoud
                # confirmed their own Net sales/COD-derived values are
                # already correct as-is, with or without a multi-row order.
                # Falls back to Total sales only when no Net sales column was
                # mapped at all (nothing to add the VAT back onto).
                if orow is None:
                    val = None
                elif orow['_net_sales'] is not None:
                    val = orow['_net_sales'] * 1.05
                else:
                    val = orow['_order_value']
            elif source == 'uae_vat_computed':
                # UAE & Oman Shipping (Sep 2026) -- Total sales minus the
                # SAME Net x1.05 figure used for Order Value above (not raw
                # Net sales), so Value + Shipping still equals Total sales,
                # nothing double-counted -- same principle as Iraq's plain
                # Total-minus-Net ('computed' above), just with the VAT
                # add-back folded into the subtraction too.
                if orow is not None and orow['_net_sales'] is not None:
                    val = orow['_order_value'] - (orow['_net_sales'] * 1.05)
                else:
                    val = None
            elif source == 'zero':
                # Gulf Shipping -- not split out at all, always written as 0
                # (no reliable source to derive it from). Order Value now
                # comes from Shopify's own Total sales instead (Sep 2026, see
                # TARGET_SHEETS['gulf']'s 'fields' comment) -- this branch is
                # unrelated to that and unchanged.
                val = 0
            elif source == 'month':
                # "Sep 2026" style label derived from Order date (Sep 2026, per
                # Mahmoud -- all 3 groups). Same _order_date the 'order_date' field
                # itself uses, so it always agrees with whatever's in that column on
                # the same row -- None (blank) when unmatched, same as order_date.
                val = format_month_year(orow['_order_date']) if orow is not None else None
            elif source == 'shipping_country':
                # Gulf's output 'Consignee City' column is actually the
                # COUNTRY code (SA/KW/QA), confirmed by Mahmoud (Aug 2026) --
                # sourced from the shipping file's own 'Consignee Country'
                # column (mapped as 'country'), NOT its 'Consignee City'
                # column (a real city name, not what this output wants).
                # Reuses the same normalize_city() UAE & Oman already uses
                # for its country-code city, and the same blank/unrecognized
                # tracking so those still get flagged and counted.
                country_col = shipping_map.get('country')
                raw_country = srow_ship.get(country_col, '') if (srow_ship is not None and country_col) else ''
                val, was_blank, was_unrec = normalize_city(raw_country, target_key, default_city_for_blank)
                if was_blank:
                    blank_city_count += 1
                if was_unrec:
                    unrecognized_city_count += 1
                    issues.append(f"unrecognized shipping country: {raw_country!r}")
            else:  # shipping
                col = shipping_map.get(field)
                raw_val = srow_ship.get(col, '') if (srow_ship is not None and col) else ''
                val = clean_display(raw_val)
                if field == 'phone':
                    # Per Mahmoud (Aug 2026): strip a leading '+' from phone
                    # numbers regardless of country -- the raw sheet stores
                    # phone numbers without it.
                    val = val.lstrip('+')
            row_out[header] = val

    out_rows = []

    if join_base == 'shopify':
        for _, orow in agg.iterrows():
            key = orow['_key']
            ship_matches = ship_index.get(key, [])
            issues = []
            srow_ship = None
            if ship_matches:
                if len(ship_matches) > 1:
                    issues.append(f'{len(ship_matches)} shipping rows share this reference -- used the first')
                srow_ship = shipping_df.loc[ship_matches[0]]
            else:
                issues.append('no matching shipping-company row yet')
            if orow.get('_date_uncertain'):
                issues.append('order date may be wrong -- earliest row in this file is not a positive sale')

            row_out = {ref_header: make_ref_display(orow['_ref_raw'])}
            fill_fields(row_out, issues, orow, srow_ship)
            row_out['_matched'] = bool(ship_matches)
            row_out['_multi_row_order'] = orow['_row_count'] > 1
            row_out['_issues'] = '; '.join(issues)
            out_rows.append(row_out)
        matched_keys = {orow['_key'] for orow in [r for _, r in agg.iterrows()] if ship_index.get(orow['_key'])}
        unmatched_other = sum(len(idxs) for k, idxs in ship_index.items() if k not in matched_keys)
        unmatched_msg = (
            f'{unmatched_other} shipping-company row(s) had no matching Shopify order -- check for a '
            f'reference-number mismatch, or these may be manual/offline orders not tracked in Shopify.'
        )
    else:  # join_base == 'shipping' (all three groups now, Aug 2026)
        seen_keys = []
        for key, idxs in ship_index.items():
            seen_keys.append(key)
            issues = []
            if len(idxs) > 1:
                issues.append(f'{len(idxs)} shipping rows share this reference -- used the first')
            srow_ship = shipping_df.loc[idxs[0]]
            orow = agg_index.get(key)
            if orow is None:
                issues.append('no matching Shopify order found yet')
            elif orow.get('_date_uncertain'):
                issues.append('order date may be wrong -- earliest row in this file is not a positive sale')

            ref_raw = srow_ship.get(ship_ref_col, '') if ship_ref_col else ''
            row_out = {ref_header: make_ref_display(ref_raw)}
            fill_fields(row_out, issues, orow, srow_ship)
            row_out['_matched'] = orow is not None
            row_out['_multi_row_order'] = bool(orow is not None and orow['_row_count'] > 1)
            row_out['_issues'] = '; '.join(issues)
            out_rows.append(row_out)
        matched_keys = set(seen_keys)
        unmatched_other = sum(1 for k in agg_index if k not in matched_keys)
        unmatched_msg = (
            f'{unmatched_other} Shopify order(s) have no matching shipping-company row yet -- they are NOT '
            f"included in this download ({TARGET_SHEETS[target_key]['label']} is joined shipping-first). "
            f'They will appear once the shipping company export includes them.'
        )

    stats['blank_city_defaulted'] = blank_city_count
    stats['unrecognized_city'] = unrecognized_city_count
    if blank_city_count:
        warnings.append(
            f"{blank_city_count} order(s) had no Shipping country on the Shopify side (in-store/POS/"
            f"Draft/Mobile orders have no shipping address) -- defaulted Consignee City to "
            f"'{default_city_for_blank}'. Check those rows before trusting the country breakdown."
        )
    if unrecognized_city_count:
        warnings.append(f"{unrecognized_city_count} order(s) had a shipping country this tool didn't recognize for {TARGET_SHEETS[target_key]['label']} -- left as typed, check the '_issues' column.")

    stats['unmatched_shipping_rows'] = unmatched_other
    if unmatched_other:
        warnings.append(unmatched_msg)

    merged = pd.DataFrame(out_rows)
    stats['output_rows'] = len(merged)
    stats.setdefault('shopify_rows_kept', len(merged))
    stats['matched'] = int(merged['_matched'].sum()) if len(merged) else 0
    return merged, warnings, stats


def workbook_to_bytes(df, target_key, include_shipping_fee=False):
    """Styled .xlsx with exactly the target sheet's header row and column
    order, ready to select-all/copy and paste into the raw Google Sheet.
    Order date is written as a real Excel date cell (not typed text), which
    sidesteps the day_first/month_first ambiguity that caused the Iraq/Gulf
    date bugs -- a real date cell carries its value unambiguously on paste."""
    fields = list(TARGET_SHEETS[target_key]['fields'])
    if include_shipping_fee and not any(f == 'shipping_fee' for _, f, _ in fields):
        fields = fields + [('Shipping', 'shipping_fee', 'computed')]
    headers = [h for h, field, source in fields]

    wb = Workbook()
    ws = wb.active
    safe_title = re.sub(r'[\[\]:*?/\\]', '-', TARGET_SHEETS[target_key]['label'])[:31]
    ws.title = safe_title

    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E3D', end_color='1F4E3D', fill_type='solid')
    body_font = Font(name='Arial')

    for c, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Some real templates repeat a blank/unnamed header more than once (Gulf
    # has two) -- selecting the same column name twice via df[[...]] turns
    # .get() into a Series instead of a scalar. Caching one value-list per
    # DISTINCT column name and writing it into every position that shares
    # that header sidesteps that entirely (both blank columns are always
    # blank anyway, so there's nothing lost by this).
    col_cache = {
        col_name: (df[col_name].tolist() if col_name in df.columns else [None] * len(df))
        for col_name in set(headers)
    }

    for r in range(len(df)):
        for c, col_name in enumerate(headers, start=1):
            val = col_cache[col_name][r]
            cell = ws.cell(row=r + 2, column=c, value=val)
            cell.font = body_font
            if isinstance(val, dt.date):
                cell.number_format = 'MM/DD/YYYY'

    for c, col_name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(col_name) + 2)

    ws.freeze_panes = 'A2'

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
